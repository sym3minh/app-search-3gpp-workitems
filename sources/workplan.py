"""
workplan.py — WorkPlan download, search, export.

Import: config
Toàn bộ logic liên quan đến file workplan.xlsx của 3GPP.
"""

import ssl, sys, os, re, datetime, subprocess
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from config import (
    CACHE_FILE, CACHE_DIR, CACHE_MAX_DAYS,
    WORKPLAN_INDEX, PORTAL_BASE,
    HDRS, OUTPUT_DIR,
    WI_OUTPUT_HEADER,
)


# ── SSL ───────────────────────────────────────────────────────────────────────

def make_ssl_ctx():
    """Tạo SSL context bỏ qua certificate verification (3GPP dùng cert tự ký)."""
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


# ── Portal check helpers ──────────────────────────────────────────────────────

def has_change_request(uid: str) -> bool:
    url = f"{PORTAL_BASE}/ChangeRequests.aspx?q=1&workitem={uid}"
    try:
        req = urllib.request.Request(url, headers=HDRS)
        with urllib.request.urlopen(req, timeout=10, context=make_ssl_ctx()) as r:
            return "No Change Request found" not in r.read().decode("utf-8", "ignore")
    except Exception:
        return True


def has_specification(uid: str) -> bool:
    url = f"{PORTAL_BASE}/Specifications.aspx?q=1&WiUid={uid}"
    try:
        req = urllib.request.Request(url, headers=HDRS)
        with urllib.request.urlopen(req, timeout=10, context=make_ssl_ctx()) as r:
            return "No records to display" not in r.read().decode("utf-8", "ignore")
    except Exception:
        return True


# ── Parallel check helpers ────────────────────────────────────────────────────

def _pcheck(items, fn, prog_fn=None):
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = {ex.submit(fn, i["uid"]): i for i in items}
        res = {}; done = [0]
        for fut in as_completed(futs):
            done[0] += 1
            res[futs[fut]["uid"]] = fut.result()
            if prog_fn:
                prog_fn(done[0], len(items))
    return res


def parallel_check_any(items, prog_fn=None):
    cr_res = {}; spec_res = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        cr_f  = {ex.submit(has_change_request, i["uid"]): ("cr",  i) for i in items}
        sp_f  = {ex.submit(has_specification,  i["uid"]): ("sp",  i) for i in items}
        all_f = {**cr_f, **sp_f}; done = [0]
        for fut in as_completed(all_f):
            done[0] += 1
            kind, item = all_f[fut]
            (cr_res if kind == "cr" else spec_res)[item["uid"]] = fut.result()
            if prog_fn:
                prog_fn(done[0], len(items) * 2)
    out = []
    for item in items:
        hc = cr_res.get(item["uid"])
        hs = spec_res.get(item["uid"])
        if hc or hs:
            if hc:
                item["cr_link"]   = f"{PORTAL_BASE}/ChangeRequests.aspx?q=1&workitem={item['uid']}"
            if hs:
                item["spec_link"] = f"{PORTAL_BASE}/Specifications.aspx?q=1&WiUid={item['uid']}"
            out.append(item)
    return out


def parallel_check_cr(items, prog_fn=None):
    res = _pcheck(items, has_change_request, prog_fn)
    out = []
    for item in items:
        if res.get(item["uid"]):
            item["cr_link"] = f"{PORTAL_BASE}/ChangeRequests.aspx?q=1&workitem={item['uid']}"
            out.append(item)
    return out


def parallel_check_spec(items, prog_fn=None):
    res = _pcheck(items, has_specification, prog_fn)
    out = []
    for item in items:
        if res.get(item["uid"]):
            item["spec_link"] = f"{PORTAL_BASE}/Specifications.aspx?q=1&WiUid={item['uid']}"
            out.append(item)
    return out


# ── Download ──────────────────────────────────────────────────────────────────

def find_xlsx_url() -> str:
    """Scrape trang FTP để tìm URL file .xlsx mới nhất."""
    req = urllib.request.Request(WORKPLAN_INDEX, headers=HDRS)
    with urllib.request.urlopen(req, timeout=20, context=make_ssl_ctx()) as r:
        html = r.read().decode("utf-8", "ignore")
    matches = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', html, re.IGNORECASE)
    if not matches:
        raise RuntimeError("Không tìm thấy file .xlsx nào trên 3GPP FTP.")
    url = matches[-1]
    return url if url.startswith("http") else "https://www.3gpp.org" + url


def download_workplan(force: bool = False, log_fn=None):
    """
    Tải workplan.xlsx từ 3GPP FTP hoặc dùng cache nếu còn mới.
    Trả về (Path, age_days).
    """
    def log(m):
        if log_fn:
            log_fn(m)

    if CACHE_FILE.exists() and not force:
        age = (datetime.date.today() -
               datetime.date.fromtimestamp(CACHE_FILE.stat().st_mtime)).days
        if age < CACHE_MAX_DAYS:
            return CACHE_FILE, age
        log(f"Cache đã {age} ngày, tự động cập nhật...")

    CACHE_DIR.mkdir(exist_ok=True)
    log("Đang tìm file Work Plan mới nhất trên 3GPP...")
    xlsx_url = find_xlsx_url()
    log(f"Tìm thấy: {xlsx_url.split('/')[-1]} — đang tải...")
    req = urllib.request.Request(xlsx_url, headers=HDRS)
    with urllib.request.urlopen(req, timeout=90, context=make_ssl_ctx()) as r:
        data = r.read()
    if len(data) < 50_000:
        raise RuntimeError("File tải về quá nhỏ.")
    CACHE_FILE.write_bytes(data)
    log(f"Đã lưu ({len(data) // 1024} KB)")
    return CACHE_FILE, 0


# ── Search ────────────────────────────────────────────────────────────────────

def search_workitems(xlsx_path, query, release_filter=None, limit=200, case_sensitive=False):
    """
    Tìm work items theo keyword (split bằng |), filter release, giới hạn limit.
    Trả về (list[dict], total_count).
    """
    import openpyxl
    wb    = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = next(
        (wb[n] for n in wb.sheetnames
         if any(k in n.lower() for k in ("wi", "work", "item", "plan"))),
        wb.worksheets[0]
    )
    rows   = sheet.iter_rows(values_only=True)
    header = None
    for row in rows:
        if row and any(str(c or "").strip() for c in row):
            header = [str(c or "").strip().lower() for c in row]
            break
    if not header:
        raise RuntimeError("Không đọc được header.")

    def col(n):
        try:
            return header.index(n.lower())
        except ValueError:
            raise RuntimeError(f"Không tìm thấy cột '{n}'.")

    ci = {k: col(k) for k in [
        "Unique_ID", "Name", "Acronym", "Release", "Start",
        "Finish", "Completion", "Status_Report", "Impacted_TSs_and_TRs",
    ]}
    keywords = [kw.strip() for kw in query.split("|") if kw.strip()]
    if not case_sensitive:
        keywords = [kw.lower() for kw in keywords]

    results = []; total_found = 0
    for row in rows:
        if not row or not any(row):
            continue

        def cell(k):
            i = ci[k]
            return "" if i >= len(row) else str(row[i] or "").strip()

        def cell_raw(k):
            i = ci[k]
            return None if i >= len(row) else row[i]

        uid = cell("Unique_ID")
        if not uid:
            continue
        try:
            if int(float(uid)) == 0:
                continue
        except (ValueError, TypeError):
            pass

        impacted = cell("Impacted_TSs_and_TRs")
        raw_comp = row[ci["Completion"]] if ci["Completion"] < len(row) else None
        try:
            pct = float(raw_comp or 0)
            if pct <= 1.0:
                pct *= 100
            comp_disp = str(int(round(pct))) + "%"
        except (TypeError, ValueError):
            comp_disp = str(raw_comp or "")
        if comp_disp in ("0%", ""):
            continue

        release = cell("Release")
        if release_filter:
            rl = release.lower()
            if all(f not in rl for f in [
                f"rel-{release_filter}", f"rel {release_filter}", release_filter
            ]):
                continue

        name = cell("Name"); code = cell("Acronym")
        nc = name if case_sensitive else name.lower()
        cc = code if case_sensitive else code.lower()
        if any(kw in nc or kw in cc for kw in keywords):
            total_found += 1
            if len(results) < limit:
                results.append({
                    "uid":      uid,
                    "code":     code,
                    "title":    name,
                    "completion": comp_disp,
                    "status":   cell("Status_Report"),
                    "release":  release,
                    "start":    cell_raw("Start"),
                    "finish":   cell_raw("Finish"),
                    "impacted": impacted,
                    "comp_raw": raw_comp,
                    "cr_link":  "",
                    "spec_link": "",
                })
    wb.close()
    return results, total_found


# ── Lookup helpers ────────────────────────────────────────────────────────────

def load_wi_by_id(uid_str: str):
    """Tra cứu 1 work item đầy đủ theo Unique_ID. Dùng bởi WI Detail tab."""
    if not CACHE_FILE.exists():
        return None
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(CACHE_FILE, read_only=True, data_only=True)
        sheet = next(
            (wb[n] for n in wb.sheetnames
             if any(k in n.lower() for k in ("wi", "work", "item", "plan"))),
            wb.worksheets[0]
        )
        rows   = sheet.iter_rows(values_only=True)
        header = None
        for row in rows:
            if row and any(str(c or "").strip() for c in row):
                header = [str(c or "").strip().lower() for c in row]
                break
        if not header:
            wb.close(); return None

        def ci(n):
            try:
                return header.index(n.lower())
            except ValueError:
                return None

        idx = {k: ci(k) for k in [
            "unique_id", "name", "acronym", "release", "start", "finish",
            "completion", "status_report", "impacted_tss_and_trs",
            "rapporteur", "tdocgroup",
        ]}
        uid_target = str(uid_str).strip()
        for row in rows:
            if not row or not any(row):
                continue
            uid_val = (
                str(row[idx["unique_id"]] or "").strip()
                if idx["unique_id"] is not None else ""
            )
            if uid_val != uid_target:
                continue

            def gv(k):
                i = idx.get(k)
                return str(row[i] or "").strip() if i is not None and i < len(row) else ""

            raw_comp = (
                row[idx["completion"]]
                if idx["completion"] is not None and idx["completion"] < len(row)
                else None
            )
            try:
                pct = float(raw_comp or 0)
                if pct <= 1.0:
                    pct *= 100
                comp_str = str(int(round(pct))) + "%"
            except Exception:
                comp_str = str(raw_comp or "")

            wb.close()
            return {
                "uid":       uid_val,
                "name":      gv("name"),
                "acronym":   gv("acronym"),
                "release":   gv("release"),
                "start":     gv("start"),
                "finish":    gv("finish"),
                "completion": comp_str,
                "status":    gv("status_report"),
                "impacted":  gv("impacted_tss_and_trs"),
                "rapporteur": gv("rapporteur"),
                "group":     gv("tdocgroup"),
            }
        wb.close()
        return None
    except Exception:
        return None


def load_workplan_wi_info(wi_ids: set) -> dict:
    """Tra cứu batch nhiều WI IDs — dùng bởi CR Search để enrich release info."""
    if not CACHE_FILE.exists():
        return {}
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(CACHE_FILE, read_only=True, data_only=True)
        sheet = next(
            (wb[n] for n in wb.sheetnames
             if any(k in n.lower() for k in ("wi", "work", "item", "plan"))),
            wb.worksheets[0]
        )
        rows   = sheet.iter_rows(values_only=True)
        header = None
        for row in rows:
            if row and any(str(c or "").strip() for c in row):
                header = [str(c or "").strip().lower() for c in row]
                break
        if not header:
            wb.close(); return {}

        def ci(n):
            try:
                return header.index(n.lower())
            except ValueError:
                return None

        idx_uid  = ci("unique_id")
        idx_name = ci("name")
        idx_rel  = ci("release")
        if idx_uid is None:
            wb.close(); return {}

        result = {}
        for row in rows:
            if not row or not any(row):
                continue
            uid = str(row[idx_uid] or "").strip()
            if uid not in wi_ids:
                continue
            name    = str(row[idx_name] or "").strip() if idx_name is not None else ""
            release = str(row[idx_rel]  or "").strip() if idx_rel  is not None else ""
            result[uid] = {"release": release, "name": name}
            if len(result) >= len(wi_ids):
                break
        wb.close()
        return result
    except Exception:
        return {}


def _load_wi_full(uid: str):
    """Load các trường bổ sung (Acronym, Completion, Status) cho Detail tab."""
    if not CACHE_FILE.exists():
        return None
    try:
        import openpyxl
        wb    = openpyxl.load_workbook(CACHE_FILE, read_only=True, data_only=True)
        sheet = next(
            (wb[n] for n in wb.sheetnames
             if any(k in n.lower() for k in ("wi", "work", "item", "plan"))),
            wb.worksheets[0]
        )
        rows   = sheet.iter_rows(values_only=True)
        header = None
        for row in rows:
            if row and any(str(c or "").strip() for c in row):
                header = [str(c or "").strip().lower() for c in row]
                break
        if not header:
            wb.close(); return None

        def ci(n):
            try:
                return header.index(n.lower())
            except Exception:
                return None

        idx = {k: ci(k) for k in [
            "Unique_ID", "Name", "Acronym", "Release", "Start",
            "Finish", "Completion", "Status_Report", "Impacted_TSs_and_TRs",
        ]}
        for row in rows:
            if not row or not any(row):
                continue
            u = str(
                row[idx["Unique_ID"]]
                if idx["Unique_ID"] is not None and idx["Unique_ID"] < len(row)
                else ""
            ).strip()
            if u != uid:
                continue

            def cell(k):
                i = idx.get(k)
                return "" if i is None or i >= len(row) else str(row[i] or "").strip()

            raw_comp = (
                row[idx["Completion"]]
                if idx["Completion"] is not None and idx["Completion"] < len(row)
                else None
            )
            try:
                pct = float(raw_comp or 0)
                if pct <= 1.0:
                    pct *= 100
                comp_str = str(int(round(pct))) + "%"
            except Exception:
                comp_str = str(raw_comp or "")

            wb.close()
            return {
                "code":       cell("Acronym"),
                "completion": comp_str,
                "status":     cell("Status_Report"),
                "start":      cell("Start"),
                "finish":     cell("Finish"),
                "impacted":   cell("Impacted_TSs_and_TRs"),
            }
        wb.close()
        return None
    except Exception:
        return None


# ── Release sort key ──────────────────────────────────────────────────────────

def _release_sort_key(rel_str: str) -> int:
    m = re.search(r'(\d+)', rel_str)
    return int(m.group(1)) if m else 0


# ── Export ────────────────────────────────────────────────────────────────────

def export_wi_xlsx(items, output_path):
    """Xuất kết quả ra Excel có style (header màu, freeze panes)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Results"
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    ws.append(WI_OUTPUT_HEADER)
    for cell in ws[1]:
        cell.font  = hfont
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for item in items:
        ws.append([
            item["uid"],   item["title"],   item["code"],  item["release"],
            item["start"], item["finish"],  item["comp_raw"], item["impacted"],
            item.get("cr_link", ""), item.get("spec_link", ""),
        ])
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0%"
    for ci, w in {1: 15, 2: 60, 3: 30, 4: 12, 5: 22, 6: 22, 7: 14, 8: 40, 9: 60, 10: 60}.items():
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)


def build_wi_filename(query, release, limit, check_cr, check_spec, check_any, case_sensitive):
    """Tạo tên file output từ query params."""
    safe  = re.sub(r'_+', '_', re.sub(r'[^a-zA-Z0-9_\-]', '_', query)).strip('_')
    parts = [safe]
    if release:       parts.append("rel" + release)
    if limit != 200:  parts.append("n"   + str(limit))
    if check_any:     parts.append("check_any")
    elif check_cr:    parts.append("check_cr")
    elif check_spec:  parts.append("check_spec")
    if case_sensitive: parts.append("cs")
    return "_".join(parts) + ".xlsx"


# ── OS open ───────────────────────────────────────────────────────────────────

def open_file(path):
    """Mở file bằng OS default app (cross-platform: win32/darwin/linux)."""
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])
    except Exception as e:
        from tkinter import messagebox
        messagebox.showerror("Lỗi", f"Không mở được:\n{e}")
