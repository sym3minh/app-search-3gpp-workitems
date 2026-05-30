"""
cr_search.py — Tìm kiếm CR Titles database (cr_titles.db).

Import: config, tdoc._detect_ext
Database được tạo bởi công cụ cr_indexer.py (ngoài project này).
"""

import re, os, sqlite3
from pathlib import Path

from config import (
    DB_FILE, PORTAL_BASE, TDOC_FETCH_OK, HDRS,
    OUTPUT_DIR, DOWNLOAD_ZIP_DIR, DOWNLOAD_EXTRACTED_DIR,
    CR_OUTPUT_HEADER_FULL, CR_OUTPUT_HEADER_WIONLY,
)
# _detect_ext được import từ tdoc — dependency một chiều an toàn
# (tdoc không import cr_search ở top-level, chỉ import lazy trong tdoc_fetch_from_db)
from tdoc import _detect_ext, _extract_zip_to, _zip_cache_path, _find_zip_in_cache


# ── DB status ─────────────────────────────────────────────────────────────────

def cr_db_status():
    """
    Kiểm tra trạng thái DB: tồn tại, kích thước, số titles, số WIs, ngày crawl.
    Trả về tuple (exists, kb, total_titles, total_workitems, last_crawled).
    """
    if not DB_FILE.exists():
        return False, 0, 0, 0, None
    try:
        conn = sqlite3.connect(str(DB_FILE))
        conn.row_factory = sqlite3.Row
        total_titles    = conn.execute("SELECT COUNT(*) FROM cr_titles").fetchone()[0]
        total_workitems = conn.execute("SELECT COUNT(*) FROM workitems").fetchone()[0]
        last_crawled    = conn.execute("SELECT MAX(last_crawled) FROM workitems").fetchone()[0]
        conn.close()
        kb = DB_FILE.stat().st_size // 1024
        return True, kb, total_titles, total_workitems, last_crawled
    except Exception:
        return True, DB_FILE.stat().st_size // 1024, 0, 0, None


# ── Search ────────────────────────────────────────────────────────────────────

def cr_search(query: str, limit: int = 100, workitem_only: bool = False):
    """
    Tìm kiếm FTS5 trên bảng cr_titles_fts.
    Fallback sang LIKE nếu FTS5 không có.
    Hỗ trợ syntax: OR, "exact phrase", keyword*

    Returns (list[dict], total_count).
    """
    if not DB_FILE.exists():
        raise RuntimeError(
            f"Chưa có database tại {DB_FILE}\nChạy cr_indexer.py trước."
        )

    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row

    # Detect schema
    cols    = {row[1] for row in conn.execute("PRAGMA table_info(cr_titles)").fetchall()}
    has_wg  = "wg_tdoc"      in cols
    has_tsg = "tsg_tdoc"     in cols
    has_dl  = "download_url" in cols

    if has_wg and has_tsg:
        extra_sel = ", cr_titles.wg_tdoc, cr_titles.tsg_tdoc"
    elif has_wg:
        extra_sel = ", cr_titles.wg_tdoc, '' AS tsg_tdoc"
    elif has_tsg:
        extra_sel = ", '' AS wg_tdoc, cr_titles.tsg_tdoc"
    elif has_dl:
        extra_sel = ", cr_titles.download_url AS wg_tdoc, '' AS tsg_tdoc"
    else:
        extra_sel = ", '' AS wg_tdoc, '' AS tsg_tdoc"

    # Build FTS5 query
    fts_query = query
    if not any(op in query for op in (" OR ", " NOT ", '"', "NEAR")):
        words     = query.split()
        fts_query = (" ".join(f'"{w}"*' for w in words)
                     if len(words) > 1 else f'"{query}"*')

    try:
        rows = conn.execute(f"""
            SELECT cr_titles.title, cr_titles.workitem_id{extra_sel}, rank
            FROM   cr_titles_fts
            JOIN   cr_titles ON cr_titles.id = cr_titles_fts.rowid
            WHERE  cr_titles_fts MATCH ?
            ORDER BY rank LIMIT ?
        """, (fts_query, limit)).fetchall()
    except sqlite3.OperationalError:
        rows = conn.execute(f"""
            SELECT title, workitem_id{extra_sel}, 0 AS rank
            FROM   cr_titles
            WHERE  title LIKE ? COLLATE NOCASE
            LIMIT  ?
        """, (f"%{query}%", limit)).fetchall()
    conn.close()

    def _build_dl_url(row):
        wg    = (row["wg_tdoc"]  or "").strip()
        tsg   = (row["tsg_tdoc"] or "").strip()
        tdoc_id = wg if wg else tsg
        if not tdoc_id:
            return ""
        if tdoc_id.startswith("http"):
            return tdoc_id
        return (f"https://portal.3gpp.org/ngppapp/DownloadTDoc.aspx"
                f"?contributionUid={tdoc_id}")

    if workitem_only:
        seen: dict = {}
        for r in rows:
            wi = str(r["workitem_id"])
            seen.setdefault(wi, []).append(r["title"])
        result = []
        for wi, titles in seen.items():
            url = f"{PORTAL_BASE}/ChangeRequests.aspx?q=1&workitem={wi}"
            result.append({"title": f"{len(titles)} title(s)", "workitem_id": wi,
                           "portal_url": url, "download_url": ""})
        return result, len(seen)
    else:
        seen_titles: set = set()
        result = []
        for r in rows:
            t = r["title"]
            if t in seen_titles:
                continue
            seen_titles.add(t)
            wi  = str(r["workitem_id"])
            url = f"{PORTAL_BASE}/ChangeRequests.aspx?q=1&workitem={wi}"
            dl  = _build_dl_url(r)
            result.append({"title": t, "workitem_id": wi, "portal_url": url, "download_url": dl})
        return result, len(result)


# ── Single file download ──────────────────────────────────────────────────────

def download_cr_file(dl_url: str, extract_dir: Path, hint_name: str = None) -> "Path | None":
    """
    Tải 1 file CR.

    ZIP  → saved to data/downloads/Zip/<LETTER>/<fname>.zip (shared cache).
           Extracted to extract_dir/<stem>/
    Non-ZIP → saved directly to extract_dir.

    Returns Path of saved file, or None on failure.
    """
    if not TDOC_FETCH_OK:
        raise RuntimeError("pip install requests")
    if not dl_url or dl_url.strip().startswith("javascript:"):
        return None

    from urllib.parse import urljoin, urlparse, parse_qs
    from bs4 import BeautifulSoup
    import requests

    if not dl_url.startswith("http"):
        dl_url = urljoin("https://portal.3gpp.org/", dl_url)

    sess = requests.Session()
    sess.headers.update({"User-Agent": HDRS["User-Agent"]})
    try:
        resp    = sess.get(dl_url, timeout=60, verify=False)
        content = resp.content

        # HTML redirect page
        if len(content) < 5000 and b'<html' in content[:200].lower():
            s = BeautifulSoup(content, 'html.parser')
            real_url = None
            for sc in s.find_all('script'):
                txt = sc.string or ''
                if 'window.location.href' in txt:
                    m = re.search(r"window\.location\.href\s*=\s*['\"]([^'\"]+)['\"]", txt)
                    if m:
                        real_url = m.group(1)
                        break
            if not real_url:
                return None
            resp    = sess.get(real_url, timeout=60, verify=False)
            content = resp.content

        if len(content) < 200 or b'<html' in content[:200].lower():
            return None

        ct  = resp.headers.get('content-type', '')
        cd  = resp.headers.get('content-disposition', '')
        ext, cd_filename = _detect_ext(content, ct, cd)

        if cd_filename:
            stem  = os.path.splitext(cd_filename)[0]
            fname = cd_filename
        else:
            qs        = parse_qs(urlparse(dl_url).query)
            uid_param = (qs.get('contributionUid') or qs.get('tdocuid') or [None])[0]
            if uid_param:
                stem = uid_param
            elif hint_name:
                stem = re.sub(r'[^\w\-]', '_', hint_name[:60]).strip('_')
            else:
                seg  = dl_url.rstrip('/').split('/')[-1].split('?')[0]
                stem = os.path.splitext(seg)[0] if '.' in seg else seg or 'cr_file'
            fname = f"{stem}{ext}"

        if ext == '.zip':
            zip_path = _zip_cache_path(fname)
            if not (zip_path.exists() and zip_path.stat().st_size > 500):
                zip_path.write_bytes(content)
            xdir = extract_dir / stem
            if not xdir.exists():
                _extract_zip_to(zip_path, xdir)
            return zip_path
        else:
            extract_dir.mkdir(parents=True, exist_ok=True)
            fp = extract_dir / fname
            if fp.exists() and fp.stat().st_size > 500:
                return fp
            fp.write_bytes(content)
            return fp

    except Exception:
        return None


# ── Export ────────────────────────────────────────────────────────────────────

def export_cr_xlsx(rows, output_path, workitem_only: bool = False):
    """Xuất kết quả tìm kiếm ra Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CR Search Results"
    hfill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont  = Font(bold=True, color="FFFFFF")
    header = CR_OUTPUT_HEADER_WIONLY if workitem_only else CR_OUTPUT_HEADER_FULL
    ws.append(header)
    for cell in ws[1]:
        cell.font  = hfont
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([row["title"], row["workitem_id"], row.get("extra", ""), row["portal_url"]])
    for ci, w in {1: 80, 2: 15, 3: 40, 4: 70}.items():
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(output_path)


def build_cr_filename(query: str, limit: int, workitem_only: bool) -> str:
    """Tạo tên file output."""
    safe  = re.sub(r'_+', '_', re.sub(r'[^a-zA-Z0-9_\-]', '_', query)).strip('_')
    parts = ["cr", safe]
    if limit != 100:    parts.append("n" + str(limit))
    if workitem_only:   parts.append("wi_only")
    return "_".join(parts) + ".xlsx"
